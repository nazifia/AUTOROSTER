import calendar
from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from datetime import timedelta

from .models import PtechStaffEntry, RosterEntry, StaffAvailability


def _build_unavailability_map(staff_ids, year, month):
    """Return {date: set(staff_id)} for unavailable staff in the given month."""
    _, num_days = calendar.monthrange(year, month)
    month_start = date(year, month, 1)
    month_end = date(year, month, num_days)

    records = StaffAvailability.objects.filter(
        staff_id__in=staff_ids,
        start_date__lte=month_end,
        end_date__gte=month_start,
    ).values_list('staff_id', 'start_date', 'end_date')

    unavailable = {}
    for staff_id, start, end in records:
        cur = max(start, month_start)
        cutoff = min(end, month_end)
        while cur <= cutoff:
            unavailable.setdefault(cur, set()).add(staff_id)
            cur += timedelta(days=1)
    return unavailable


def _resolve_active_days(days_pattern, custom_days):
    """Return set of weekday ints (0=Mon … 6=Sun) based on pattern choice."""
    if days_pattern == 'weekdays':
        return {0, 1, 2, 3, 4}
    if days_pattern == 'weekends':
        return {5, 6}
    if days_pattern == 'custom' and custom_days:
        return {int(d) for d in custom_days}
    return {0, 1, 2, 3, 4, 5, 6}


def _pick(pool, idx, unavailable_ids, recently_assigned, current_day, min_gap, mode,
          quota=None, used_counts=None):
    """Pick next available staff; return (staff_or_None, new_idx).

    quota: {staff_id: max appearances}. Staff at or over their cap are skipped.
    """
    if not pool:
        return None, idx

    def eligible(s):
        if s.id in unavailable_ids:
            return False
        if quota and used_counts is not None:
            cap = quota.get(s.id)
            if cap is not None and used_counts.get(s.id, 0) >= cap:
                return False
        if min_gap and min_gap > 0:
            last = recently_assigned.get(s.id)
            if last is not None and (current_day - last) < min_gap:
                return False
        return True

    if mode == 'fixed':
        for s in pool:
            if eligible(s):
                return s, idx
        return None, idx

    for offset in range(len(pool)):
        candidate = pool[(idx + offset) % len(pool)]
        if eligible(candidate):
            return candidate, (idx + offset + 1) % len(pool)
    return None, idx


def generate_roster_entries(roster, slot1_staff, slot2_staff, slot3_staff,
                            slot1_mode='rotate', slot2_mode='rotate', slot3_mode='rotate',
                            slot1_days_pattern='all', slot1_custom_days=None, slot1_min_gap=0,
                            slot2_days_pattern='all', slot2_custom_days=None, slot2_min_gap=0,
                            slot3_days_pattern='all', slot3_custom_days=None, slot3_min_gap=0,
                            slot1_quota=None, staff_start_dates=None):
    """
    Generate RosterEntry objects for every day in the roster's month/year.
    Respects StaffAvailability, days patterns, and min-gap constraints.

    staff_start_dates: {staff_id: date}. A staff is excluded on any day before
    their start date (e.g. joined the unit mid-month).
    """
    staff_start_dates = staff_start_dates or {}
    RosterEntry.objects.filter(roster=roster).delete()

    year, month = roster.year, roster.month
    _, num_days = calendar.monthrange(year, month)

    s1_pool = list(slot1_staff)
    s2_pool = list(slot2_staff)
    s3_pool = list(slot3_staff)

    all_staff_ids = {s.id for s in s1_pool + s2_pool + s3_pool}
    unavailability = _build_unavailability_map(all_staff_ids, year, month)

    s1_active = _resolve_active_days(slot1_days_pattern, slot1_custom_days)
    s2_active = _resolve_active_days(slot2_days_pattern, slot2_custom_days)
    s3_active = _resolve_active_days(slot3_days_pattern, slot3_custom_days)

    s1_min_gap = int(slot1_min_gap or 0)
    s2_min_gap = int(slot2_min_gap or 0)
    s3_min_gap = int(slot3_min_gap or 0)

    s1_quota = slot1_quota or {}
    s1_used = {}

    s1_idx = s2_idx = s3_idx = 0
    s1_last = {}
    s2_last = {}
    s3_last = {}
    entries = []

    for day in range(1, num_days + 1):
        d = date(year, month, day)
        unavail = unavailability.get(d, set())
        not_yet = {sid for sid, sd in staff_start_dates.items() if d < sd}
        if not_yet:
            unavail = unavail | not_yet
        weekday = d.weekday()

        s1 = s2 = s3 = None

        if s1_pool and weekday in s1_active:
            s1, s1_idx = _pick(s1_pool, s1_idx, unavail, s1_last, day, s1_min_gap, slot1_mode,
                               quota=s1_quota, used_counts=s1_used)
            if s1:
                s1_last[s1.id] = day
                s1_used[s1.id] = s1_used.get(s1.id, 0) + 1

        if s2_pool and roster.num_slots >= 2 and weekday in s2_active:
            s2, s2_idx = _pick(s2_pool, s2_idx, unavail, s2_last, day, s2_min_gap, slot2_mode)
            if s2:
                s2_last[s2.id] = day

        if s3_pool and roster.num_slots >= 3 and weekday in s3_active:
            s3, s3_idx = _pick(s3_pool, s3_idx, unavail, s3_last, day, s3_min_gap, slot3_mode)
            if s3:
                s3_last[s3.id] = day

        entries.append(RosterEntry(roster=roster, date=d, slot1=s1, slot2=s2, slot3=s3))

    RosterEntry.objects.bulk_create(entries)


def _ensure_daily_coverage(days_map, required_shifts=('M', 'A', 'N')):
    """
    Post-processing pass: guarantee each shift in required_shifts has ≥1 staff every day.
    Staff on 'O' (off) are reassigned to cover missing shifts.
    Tracks reassignment counts per staff for fairness (least-reassigned picked first).
    Staff on 'L' (leave) are never reassigned.
    """
    reassign_count = {}

    for cur in sorted(days_map.keys()):
        day_list = days_map[cur]
        covered = {e['shift'] for e in day_list if e['shift'] not in ('O', 'L')}
        missing = [s for s in required_shifts if s not in covered]
        if not missing:
            continue

        off_pool = [e for e in day_list if e['shift'] == 'O']
        off_pool.sort(key=lambda e: reassign_count.get(e['staff'].id, 0))

        for needed_shift in missing:
            if not off_pool:
                break
            chosen = off_pool.pop(0)
            chosen['shift'] = needed_shift
            reassign_count[chosen['staff'].id] = reassign_count.get(chosen['staff'].id, 0) + 1


def generate_ptech_roster_entries(roster, morning_staff, afternoon_staff, cm_staff,
                                   post_cm_rest_days=2, rotate_shifts=True, cm_min_gap=0,
                                   morning_work_days=5, morning_off_days=2,
                                   afternoon_work_days=5, afternoon_off_days=2,
                                   night_work_days=2, night_off_days=5,
                                   active_shifts=None, staff_start_dates=None):
    """
    Generate PtechStaffEntry records (staff×day matrix) for a PTech roster.

    active_shifts: iterable of shift codes to include, e.g. ['M', 'A']. Defaults to all.
    rotate_shifts=True  – all staff rotate through selected shifts sequentially, staggered.
    rotate_shifts=False – staff stay in their assigned pool with work/off cycling.
    staff_start_dates: {staff_id: date}. Staff get no row on days before their
    start date (joined mid-month) – those cells stay blank.
    """
    PtechStaffEntry.objects.filter(roster=roster).delete()
    staff_start_dates = staff_start_dates or {}

    if active_shifts is None:
        active_shifts = {'M', 'A', 'N'}
    else:
        active_shifts = set(active_shifts)

    year, month = roster.year, roster.month
    _, num_days = calendar.monthrange(year, month)

    # Deduplicate, preserve order: morning → afternoon → night extras
    seen = set()
    all_staff = []
    for s in list(morning_staff) + list(afternoon_staff) + list(cm_staff):
        if s.id not in seen:
            seen.add(s.id)
            all_staff.append(s)

    if not all_staff:
        return

    unavailability = _build_unavailability_map([s.id for s in all_staff], year, month)

    # Build ordered cycle segments from active shifts only
    _shift_params = [
        ('M', morning_work_days, morning_off_days),
        ('A', afternoon_work_days, afternoon_off_days),
        ('N', night_work_days, night_off_days),
    ]
    segments = [(code, w, o) for code, w, o in _shift_params if code in active_shifts]

    # days_map: date → list of {staff, shift} dicts (mutated in place by coverage pass)
    days_map = {}

    if rotate_shifts:
        total_cycle = sum(w + o for _, w, o in segments)
        if total_cycle == 0:
            return

        n = len(all_staff)
        stagger = total_cycle // n if n > 0 else 0

        staff_phase = {s.id: (i * stagger) % total_cycle for i, s in enumerate(all_staff)}

        def _rotation_shift(day_1idx, phase):
            pos = (day_1idx - 1 + phase) % total_cycle
            for code, work, off in segments:
                if pos < work:
                    return code
                pos -= work
                if pos < off:
                    return 'O'
                pos -= off
            return 'O'

        for d in range(1, num_days + 1):
            cur = date(year, month, d)
            unavail = unavailability.get(cur, set())
            days_map[cur] = [
                {'staff': staff, 'shift': 'L' if staff.id in unavail else _rotation_shift(d, staff_phase[staff.id])}
                for staff in all_staff
                if not (staff_start_dates.get(staff.id) and cur < staff_start_dates[staff.id])
            ]

    else:
        # Fixed-pool mode: each staff stays in their assigned shift pool
        morning_id_set = {s.id for s in morning_staff}
        afternoon_id_set = {s.id for s in afternoon_staff}
        night_id_set = {s.id for s in cm_staff}

        def _cycle(work, off):
            return (work + off) if work > 0 else 0

        m_cycle = _cycle(morning_work_days, morning_off_days)
        a_cycle = _cycle(afternoon_work_days, afternoon_off_days)
        n_cycle = _cycle(night_work_days, night_off_days)

        m_phase = {s.id: (i * morning_work_days) % m_cycle if m_cycle > 0 else 0
                   for i, s in enumerate(morning_staff)}
        a_phase = {s.id: (i * afternoon_work_days) % a_cycle if a_cycle > 0 else 0
                   for i, s in enumerate(afternoon_staff)}
        n_phase = {s.id: (i * night_work_days) % n_cycle if n_cycle > 0 else 0
                   for i, s in enumerate(cm_staff)}

        def _in_work_phase(day_1idx, phase, work_days, cycle):
            if cycle <= 0 or work_days <= 0:
                return True
            return ((day_1idx - 1 + phase) % cycle) < work_days

        for d in range(1, num_days + 1):
            cur = date(year, month, d)
            unavail = unavailability.get(cur, set())
            day_list = []

            for staff in all_staff:
                sd = staff_start_dates.get(staff.id)
                if sd and cur < sd:
                    continue  # not yet joined – no row this day
                if staff.id in unavail:
                    shift = 'L'
                elif staff.id in night_id_set and 'N' in active_shifts:
                    if _in_work_phase(d, n_phase[staff.id], night_work_days, n_cycle):
                        shift = 'N'
                    else:
                        if staff.id in morning_id_set and 'M' in active_shifts:
                            shift = 'M' if _in_work_phase(d, m_phase[staff.id], morning_work_days, m_cycle) else 'O'
                        elif staff.id in afternoon_id_set and 'A' in active_shifts:
                            shift = 'A' if _in_work_phase(d, a_phase[staff.id], afternoon_work_days, a_cycle) else 'O'
                        else:
                            shift = 'O'
                elif staff.id in morning_id_set and 'M' in active_shifts:
                    shift = 'M' if _in_work_phase(d, m_phase[staff.id], morning_work_days, m_cycle) else 'O'
                elif staff.id in afternoon_id_set and 'A' in active_shifts:
                    shift = 'A' if _in_work_phase(d, a_phase[staff.id], afternoon_work_days, a_cycle) else 'O'
                else:
                    shift = 'O'

                day_list.append({'staff': staff, 'shift': shift})
            days_map[cur] = day_list

    required_shifts = tuple(code for code, _, _ in segments)
    _ensure_daily_coverage(days_map, required_shifts=required_shifts)

    entries = []
    for cur, day_list in days_map.items():
        for item in day_list:
            entries.append(PtechStaffEntry(roster=roster, staff=item['staff'], date=cur, shift=item['shift']))

    PtechStaffEntry.objects.bulk_create(entries)


def export_ptech_to_excel(roster):
    """Return BytesIO of xlsx matching the FTH Katsina PTech matrix format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{roster.get_month_display()} {roster.year}"

    unit = roster.unit
    dept = unit.department
    year, month_num = roster.year, roster.month
    _, num_days = calendar.monthrange(year, month_num)
    all_dates = [date(year, month_num, d) for d in range(1, num_days + 1)]

    # Fetch all entries
    entries = roster.ptech_entries.select_related('staff').order_by('staff__name', 'date')
    # Build matrix: staff_id → {date: shift}
    staff_objs = {}  # staff_id → Staff (ordered by name via queryset)
    matrix = {}
    for e in entries:
        staff_objs[e.staff_id] = e.staff
        matrix.setdefault(e.staff_id, {})[e.date] = e.shift

    staff_list = list(staff_objs.values())

    # Styles
    thin = Side(style='thin')
    medium = Side(style='medium')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_med = Border(left=medium, right=medium, top=medium, bottom=medium)

    header_fill = PatternFill('solid', fgColor='1B4F72')
    sub_fill = PatternFill('solid', fgColor='154360')
    date_hdr_fill = PatternFill('solid', fgColor='2E86C1')
    weekend_fill = PatternFill('solid', fgColor='FEF9E7')
    m_fill = PatternFill('solid', fgColor='D6EAF8')
    a_fill = PatternFill('solid', fgColor='D5F5E3')
    n_fill = PatternFill('solid', fgColor='D7BDE2')
    cm_fill = PatternFill('solid', fgColor='FDEBD0')
    l_fill = PatternFill('solid', fgColor='FADBD8')
    o_fill = PatternFill('solid', fgColor='F2F3F4')

    shift_fills = {'M': m_fill, 'A': a_fill, 'N': n_fill, 'CM': cm_fill, 'L': l_fill, 'O': o_fill}

    # Split into two halves (days 1-16 and 17-end) like the reference
    halves = [all_dates[:16], all_dates[16:]]

    # Header rows (spanning full width: name col + up to 16 day cols + 1 spacer = 18 cols)
    max_cols = 18

    def merge_header(row, val, fill, font_size=12, bold=True, color='FFFFFF'):
        end_col = get_column_letter(max_cols)
        ws.merge_cells(f'A{row}:{end_col}{row}')
        cell = ws[f'A{row}']
        cell.value = val
        cell.font = Font(name='Calibri', bold=bold, size=font_size, color=color)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_med
        ws.row_dimensions[row].height = 22

    merge_header(1, dept.hospital.name.upper(), header_fill, font_size=14)
    merge_header(2, dept.department_name.upper(), header_fill, font_size=12)
    merge_header(3, unit.unit_name.upper(), header_fill, font_size=11)
    merge_header(4, roster.roster_title.upper(), sub_fill, font_size=13)
    merge_header(5, roster.month_year_display, PatternFill('solid', fgColor='1A5276'), font_size=12)

    # Name column width
    ws.column_dimensions['A'].width = 26
    for col_idx in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 5

    start_row = 6
    for half_dates in halves:
        if not half_dates:
            continue

        date_row = start_row
        day_row = start_row + 1
        data_start = start_row + 2

        # DATE header row
        ws.row_dimensions[date_row].height = 16
        cell = ws.cell(row=date_row, column=1, value='')
        cell.fill = date_hdr_fill
        cell.border = border_all
        cell = ws.cell(row=date_row, column=2, value='DATE')
        cell.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
        cell.fill = date_hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_all
        for ci, d in enumerate(half_dates, start=3):
            cell = ws.cell(row=date_row, column=ci, value=d.day)
            cell.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
            cell.fill = date_hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_all

        # DAY header row
        ws.row_dimensions[day_row].height = 16
        cell = ws.cell(row=day_row, column=1, value='')
        cell.fill = date_hdr_fill
        cell.border = border_all
        cell = ws.cell(row=day_row, column=2, value='DAY')
        cell.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
        cell.fill = date_hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_all
        for ci, d in enumerate(half_dates, start=3):
            is_wknd = d.weekday() >= 5
            cell = ws.cell(row=day_row, column=ci, value=d.strftime('%a').upper())
            cell.font = Font(name='Calibri', bold=True, size=9,
                             color='C0392B' if is_wknd else 'FFFFFF')
            cell.fill = date_hdr_fill if not is_wknd else weekend_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_all

        # Staff rows
        for ri, staff in enumerate(staff_list):
            row = data_start + ri
            ws.row_dimensions[row].height = 17
            cell = ws.cell(row=row, column=1, value=staff.display_name)
            cell.font = Font(name='Calibri', bold=True, size=9, color='1B2631')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border_all

            cell = ws.cell(row=row, column=2, value='')
            cell.border = border_all

            for ci, d in enumerate(half_dates, start=3):
                shift = matrix.get(staff.id, {}).get(d, '')
                is_wknd = d.weekday() >= 5
                fill = shift_fills.get(shift, weekend_fill if is_wknd else PatternFill())
                cell = ws.cell(row=row, column=ci, value=shift)
                cell.font = Font(name='Calibri', bold=(shift == 'CM'), size=9,
                                 color='7D6608' if is_wknd else '1B2631')
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_all

        start_row = data_start + len(staff_list) + 1  # blank row between halves

    # ── Staff Directory ───────────────────────────────────────────────────────
    dir_start = start_row + 1
    dir_fill = PatternFill('solid', fgColor='1B4F72')
    dir_hdr_fill = PatternFill('solid', fgColor='2E86C1')
    dir_alt_fill = PatternFill('solid', fgColor='EBF5FB')
    dir_white_fill = PatternFill('solid', fgColor='FFFFFF')

    ws.merge_cells(f'A{dir_start}:C{dir_start}')
    cell = ws[f'A{dir_start}']
    cell.value = 'STAFF DIRECTORY'
    cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    cell.fill = dir_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_med
    ws.row_dimensions[dir_start].height = 20

    hdr_row = dir_start + 1
    for col, label in [(1, 'S/N'), (2, 'NAME'), (3, 'PHONE')]:
        cell = ws.cell(row=hdr_row, column=col, value=label)
        cell.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
        cell.fill = dir_hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_all
    ws.row_dimensions[hdr_row].height = 16
    ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width, 30)
    ws.column_dimensions['C'].width = 18

    for i, staff in enumerate(staff_list, start=1):
        row = hdr_row + i
        row_fill = dir_alt_fill if i % 2 == 0 else dir_white_fill
        ws.row_dimensions[row].height = 16
        for col, val in [(1, i), (2, staff.display_name), (3, staff.phone_number or '')]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(name='Calibri', size=9, color='1B2631')
            cell.fill = row_fill
            cell.border = border_all
            cell.alignment = Alignment(horizontal='center' if col != 2 else 'left', vertical='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_roster_to_excel(roster):
    """Return BytesIO of xlsx matching the FTH Katsina roster format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{roster.get_month_display()} {roster.year}"

    unit = roster.unit
    dept = unit.department
    num_slots = roster.num_slots

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 4
    ws.column_dimensions['G'].width = 28

    total_cols = 3 + (num_slots - 1) * 2  # 3 for 1 slot, 5 for 2, 7 for 3
    last_col = get_column_letter(max(total_cols, 3))

    thin = Side(style='thin')
    medium = Side(style='medium')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)

    header_fill = PatternFill('solid', fgColor='1B4F72')
    col_fill = PatternFill('solid', fgColor='2E86C1')
    alt_fill = PatternFill('solid', fgColor='EBF5FB')
    weekend_fill = PatternFill('solid', fgColor='FEF9E7')
    white_fill = PatternFill('solid', fgColor='FFFFFF')

    def merge_write(row, val, fill, font_size=12, bold=True, color='FFFFFF'):
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws[f'A{row}']
        cell.value = val
        cell.font = Font(name='Calibri', bold=bold, size=font_size, color=color)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_medium
        ws.row_dimensions[row].height = 22

    # ── Header rows ──────────────────────────────────────────────────────────
    merge_write(1, dept.hospital.name.upper(), header_fill, font_size=14)
    merge_write(2, dept.department_name.upper(), header_fill, font_size=12)
    merge_write(3, unit.unit_name.upper(), header_fill, font_size=11)
    merge_write(4, roster.roster_title.upper(), PatternFill('solid', fgColor='154360'), font_size=13)
    merge_write(5, roster.month_year_display, PatternFill('solid', fgColor='1A5276'), font_size=12)

    # ── Column headers ────────────────────────────────────────────────────────
    ws.row_dimensions[6].height = 20
    col_headers = ['DAYS', 'DATE', roster.slot1_label]
    if num_slots >= 2:
        col_headers += ['', roster.slot2_label]
    if num_slots >= 3:
        col_headers += ['', roster.slot3_label]

    for col_idx, hdr in enumerate(col_headers, start=1):
        cell = ws.cell(row=6, column=col_idx, value=hdr)
        cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        cell.fill = col_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_all

    # ── Data rows ─────────────────────────────────────────────────────────────
    entries = roster.entries.select_related('slot1', 'slot2', 'slot3').order_by('date')

    last_data_row = 6
    for row_idx, entry in enumerate(entries, start=7):
        last_data_row = row_idx
        ws.row_dimensions[row_idx].height = 18
        is_weekend = entry.date.weekday() in (5, 6)
        row_fill = weekend_fill if is_weekend else (alt_fill if row_idx % 2 == 0 else white_fill)

        data = [
            entry.day_abbr,
            entry.date_display,
            entry.slot1.display_name if entry.slot1 else '',
        ]
        if num_slots >= 2:
            data += ['', entry.slot2.display_name if entry.slot2 else '']
        if num_slots >= 3:
            data += ['', entry.slot3.display_name if entry.slot3 else '']

        for col_idx, val in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = border_all
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=10,
                             bold=(col_idx == 1),
                             color='7D6608' if is_weekend else '1B2631')

    # ── Staff Directory ───────────────────────────────────────────────────────
    all_slot_staff = {}
    for entry in roster.entries.select_related('slot1', 'slot2', 'slot3'):
        for s in [entry.slot1, entry.slot2, entry.slot3]:
            if s and s.id not in all_slot_staff:
                all_slot_staff[s.id] = s
    dir_staff = sorted(all_slot_staff.values(), key=lambda s: s.name)

    dir_start = last_data_row + 2
    dir_fill = PatternFill('solid', fgColor='1B4F72')
    dir_hdr_fill = PatternFill('solid', fgColor='2E86C1')
    dir_alt_fill = PatternFill('solid', fgColor='EBF5FB')
    dir_white_fill = PatternFill('solid', fgColor='FFFFFF')

    ws.merge_cells(f'A{dir_start}:C{dir_start}')
    cell = ws[f'A{dir_start}']
    cell.value = 'STAFF DIRECTORY'
    cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    cell.fill = dir_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_medium
    ws.row_dimensions[dir_start].height = 20

    hdr_row = dir_start + 1
    for col, label in [(1, 'S/N'), (2, 'NAME'), (3, 'PHONE')]:
        cell = ws.cell(row=hdr_row, column=col, value=label)
        cell.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
        cell.fill = dir_hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_all
    ws.row_dimensions[hdr_row].height = 16

    for i, staff in enumerate(dir_staff, start=1):
        row = hdr_row + i
        row_fill = dir_alt_fill if i % 2 == 0 else dir_white_fill
        ws.row_dimensions[row].height = 16
        for col, val in [(1, i), (2, staff.display_name), (3, staff.phone_number or '')]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(name='Calibri', size=9, color='1B2631')
            cell.fill = row_fill
            cell.border = border_all
            cell.alignment = Alignment(horizontal='center' if col != 2 else 'left', vertical='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
